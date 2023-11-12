import openpyxl


wb = openpyxl.load_workbook("C:\\Users\\User\\PycharmProjects\\pythonProject1\\ExcelFiles\\Book.xlsx")
# suites_sheet = wb["Sheet1"]
active_sheet = wb.active
# print(suites_sheet.max_row)
# print(suites_sheet.max_column)
# print(wb.active)
column_names = [cell.value for cell in active_sheet[1]]
# if column_names == "Username":

print(column_names)

# for i in range(1, suites_sheet.max_row+1):
#     for j in range(1, suites_sheet.max_column+1):
#         print("display")
        # print(suites_sheet.cell(i,j).value)

# def getDataBasedOnColumn(name):
#     for j in range(1, suites_sheet.max_column + 1):
#         columns = suites_sheet.columns
#         # if columns